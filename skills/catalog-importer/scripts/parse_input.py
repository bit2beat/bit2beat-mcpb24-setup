#!/usr/bin/env python3
"""
parse_input.py — Normaliza archivos Excel/CSV para catalog-importer.

Uso:
    python parse_input.py --file ruta/al/archivo.xlsx [--output normalized_data.json]
    python parse_input.py --file ruta/al/archivo.csv [--sheet "Productos"]

Salida:
    JSON con detected_sheets, normalized rows por tipo, y warnings.
"""

import argparse
import json
import os
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Aliases: mapeo de nombres de columna heterogéneos → campo canónico
# ---------------------------------------------------------------------------
COLUMN_ALIASES = {
    "sku":        ["codigo", "code", "sku", "cod", "ref", "referencia", "item_code",
                   "artículo", "articulo", "part_number", "pn", "id_producto"],
    "name":       ["nombre", "name", "descripcion", "descripción", "producto",
                   "articulo", "artículo", "item", "title", "titulo", "título"],
    "section":    ["seccion", "sección", "categoria", "categoría", "rubro",
                   "familia", "category", "grupo", "linea", "línea", "department"],
    "section_parent": ["seccion_padre", "categoria_padre", "parent_section",
                       "parent_category", "rubro_padre"],
    "description": ["detalle", "detail", "descripcion_larga", "long_description",
                    "obs", "observacion", "observación", "notes"],
    "barcode":    ["codigo_barras", "barcode", "ean", "ean13", "upc", "gtin"],
    "price":      ["precio", "price", "pvp", "valor", "importe", "monto",
                   "precio_venta", "sale_price", "amount"],
    "price_type": ["tipo_precio", "lista", "pricelist", "price_type",
                   "tipo_lista", "nivel_precio"],
    "currency":   ["moneda", "currency", "divisa", "cur"],
    "store":      ["deposito", "depósito", "almacen", "almacén", "stock",
                   "store", "location", "ubicacion", "ubicación", "bodega"],
    "quantity":   ["cantidad", "qty", "stock", "existencia", "unidades",
                   "existencias", "inventory", "disponible"],
    "min_qty":    ["cant_minima", "cantidad_minima", "min_qty", "minimo",
                   "mínimo", "min_order", "pedido_minimo"],
    "discount_pct": ["descuento", "discount", "dto", "bonif", "bonificacion",
                     "bonificación", "rebaja", "pct_descuento"],
    "discount_from_qty": ["desde_cantidad", "from_qty", "cantidad_desde",
                           "cant_desde", "a_partir_de"],
    "active":     ["activo", "active", "habilitado", "enabled", "vigente"],
    "weight":     ["peso", "weight", "gr", "kg", "gramos", "kilogramos"],
    "measure":    ["unidad_medida", "medida", "measure", "unit", "um"],
    # Variantes
    "variant_property": ["propiedad", "property", "atributo", "attribute",
                          "caracteristica", "característica"],
    "variant_value":    ["valor_propiedad", "valor_atributo", "value",
                          "talle", "color", "tamaño", "tamano", "size"],
}

# Invertir el diccionario para lookup rápido: alias → canónico
_ALIAS_MAP = {}
for canonical, aliases in COLUMN_ALIASES.items():
    for alias in aliases:
        _ALIAS_MAP[alias.lower().strip()] = canonical

# ---------------------------------------------------------------------------
# Detección de tipo de hoja por combinación de columnas presentes
# ---------------------------------------------------------------------------
def detect_sheet_type(canonical_cols):
    """
    Retorna uno de: 'products', 'prices', 'stock', 'discounts', 'mixed', 'unknown'.
    'mixed' significa que contiene columnas de múltiples tipos (lo más común en
    archivos de un solo sheet).
    """
    has_product = bool({"sku", "name"} & canonical_cols)
    has_price   = bool({"price"} & canonical_cols)
    has_stock   = bool({"quantity", "store"} & canonical_cols)
    has_discount = bool({"discount_pct", "discount_from_qty"} & canonical_cols)

    types = []
    if has_product:
        types.append("products")
    if has_price:
        types.append("prices")
    if has_stock:
        types.append("stock")
    if has_discount:
        types.append("discounts")

    if len(types) == 0:
        return "unknown"
    if len(types) == 1:
        return types[0]
    return "mixed"


def normalize_column_name(raw_name):
    """Devuelve (canonical_name, is_known). Si no reconoce, devuelve (raw_name, False)."""
    key = str(raw_name).lower().strip().replace(" ", "_").replace("-", "_")
    canonical = _ALIAS_MAP.get(key)
    if canonical:
        return canonical, True
    return raw_name, False


# ---------------------------------------------------------------------------
# Parseo de Excel
# ---------------------------------------------------------------------------
def parse_excel(filepath, target_sheet=None):
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl no instalado. Ejecutar: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_to_process = [target_sheet] if target_sheet else wb.sheetnames

    result = {"detected_sheets": [], "data": {}, "warnings": []}

    for sheet_name in sheets_to_process:
        if sheet_name not in wb.sheetnames:
            result["warnings"].append(f"Hoja '{sheet_name}' no encontrada en el archivo.")
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result["warnings"].append(f"Hoja '{sheet_name}' vacía, ignorada.")
            continue

        # Primera fila = encabezados
        raw_headers = [str(h).strip() if h is not None else f"col_{i}"
                       for i, h in enumerate(rows[0])]

        col_map = {}       # raw → canonical
        unknown_cols = []
        for raw in raw_headers:
            canonical, known = normalize_column_name(raw)
            col_map[raw] = canonical
            if not known:
                unknown_cols.append(raw)

        canonical_set = set(col_map.values())
        sheet_type = detect_sheet_type(canonical_set)

        if unknown_cols:
            result["warnings"].append(
                f"Hoja '{sheet_name}': columnas no reconocidas ignoradas: {', '.join(unknown_cols)}"
            )

        # Normalizar filas de datos
        normalized_rows = []
        for row_idx, row in enumerate(rows[1:], start=2):
            if all(v is None or str(v).strip() == "" for v in row):
                continue  # fila vacía
            record = {}
            for raw_header, value in zip(raw_headers, row):
                canonical = col_map[raw_header]
                if canonical in COLUMN_ALIASES:  # solo guardar campos conocidos
                    if value is not None and str(value).strip() != "":
                        record[canonical] = str(value).strip() if isinstance(value, str) else value
            record["_source_row"] = row_idx
            normalized_rows.append(record)

        result["detected_sheets"].append({
            "sheet": sheet_name,
            "type": sheet_type,
            "rows": len(normalized_rows),
            "columns": list(canonical_set & set(COLUMN_ALIASES.keys())),
        })
        result["data"][sheet_name] = {
            "type": sheet_type,
            "rows": normalized_rows,
        }

    return result


# ---------------------------------------------------------------------------
# Parseo de CSV
# ---------------------------------------------------------------------------
def parse_csv(filepath):
    import csv
    import io

    # Detectar encoding
    try:
        with open(filepath, encoding="utf-8-sig") as f:
            content = f.read()
    except UnicodeDecodeError:
        with open(filepath, encoding="latin-1") as f:
            content = f.read()

    # Detectar delimitador
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(content[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # fallback a coma

    reader = csv.DictReader(io.StringIO(content), dialect=dialect)
    raw_headers = reader.fieldnames or []

    col_map = {}
    unknown_cols = []
    for raw in raw_headers:
        canonical, known = normalize_column_name(raw)
        col_map[raw] = canonical
        if not known:
            unknown_cols.append(raw)

    canonical_set = set(col_map.values())
    sheet_type = detect_sheet_type(canonical_set)

    warnings = []
    if unknown_cols:
        warnings.append(f"Columnas no reconocidas ignoradas: {', '.join(unknown_cols)}")

    normalized_rows = []
    for row_idx, row in enumerate(reader, start=2):
        record = {}
        for raw_header, value in row.items():
            canonical = col_map.get(raw_header, raw_header)
            if canonical in COLUMN_ALIASES and value and value.strip():
                record[canonical] = value.strip()
        if record:
            record["_source_row"] = row_idx
            normalized_rows.append(record)

    sheet_name = Path(filepath).stem
    return {
        "detected_sheets": [{
            "sheet": sheet_name,
            "type": sheet_type,
            "rows": len(normalized_rows),
            "columns": list(canonical_set & set(COLUMN_ALIASES.keys())),
        }],
        "data": {
            sheet_name: {
                "type": sheet_type,
                "rows": normalized_rows,
            }
        },
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Normaliza archivo Excel/CSV para catalog-importer.")
    parser.add_argument("--file", required=True, help="Ruta al archivo .xlsx, .xls o .csv")
    parser.add_argument("--sheet", help="Nombre de hoja específica a procesar (solo Excel)")
    parser.add_argument("--output", help="Archivo de salida JSON (default: stdout)")
    args = parser.parse_args()

    filepath = args.file
    if not os.path.exists(filepath):
        print(f"ERROR: Archivo no encontrado: {filepath}", file=sys.stderr)
        sys.exit(1)

    ext = Path(filepath).suffix.lower()
    if ext in (".xlsx", ".xls"):
        result = parse_excel(filepath, target_sheet=args.sheet)
    elif ext == ".csv":
        result = parse_csv(filepath)
    else:
        print(f"ERROR: Formato no soportado '{ext}'. Usar .xlsx, .xls o .csv", file=sys.stderr)
        sys.exit(1)

    output_json = json.dumps(result, ensure_ascii=False, indent=2, default=str)

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(output_json)
        print(f"Datos normalizados guardados en: {args.output}")
        # Imprimir resumen a stdout para que Claude lo lea
        print("\n=== RESUMEN DE DETECCIÓN ===")
        for sheet_info in result["detected_sheets"]:
            print(f"  Hoja: {sheet_info['sheet']}")
            print(f"    Tipo detectado: {sheet_info['type']}")
            print(f"    Filas: {sheet_info['rows']}")
            print(f"    Columnas reconocidas: {', '.join(sheet_info['columns'])}")
        if result["warnings"]:
            print("\nADVERTENCIAS:")
            for w in result["warnings"]:
                print(f"  - {w}")
    else:
        print(output_json)


if __name__ == "__main__":
    main()
